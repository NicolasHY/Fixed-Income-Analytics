"""
Quarterly Excel report generator.

Public API
----------
get_available_quarters(pnl)      -> list[dict]
generate_quarterly_report(quarter) -> bytes
"""
from __future__ import annotations

import io
import warnings
from datetime import date, datetime
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

TEMPLATE_PATH = Path(__file__).parent.parent / "Report_Template.xlsx"
DATA_OUT = Path(__file__).parent.parent / "data" / "output"

# ── Colour palette (matches dashboard CSS tokens) ────────────────────────────
_NAVY   = "0D1B2A"
_NAVY_M = "1B3A5C"
_GRAY_L = "E2E8F0"
_ALT    = "F6F8FB"
_WHITE  = "FFFFFF"
_TEXT   = "0F172A"
_MUTED  = "94A3B8"

# ── Style factories ───────────────────────────────────────────────────────────

def _hfill() -> PatternFill:
    return PatternFill("solid", fgColor=_NAVY)

def _sfill() -> PatternFill:
    return PatternFill("solid", fgColor=_NAVY_M)

def _cfill() -> PatternFill:
    return PatternFill("solid", fgColor=_GRAY_L)

def _rowfill(alt: bool) -> PatternFill:
    return PatternFill("solid", fgColor=_ALT if alt else _WHITE)

def _border() -> Border:
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def _hfont(size: int = 13) -> Font:
    return Font(name="Calibri", bold=True, size=size, color=_WHITE)

def _sfont() -> Font:
    return Font(name="Calibri", bold=True, size=10, color=_WHITE)

def _cfont() -> Font:
    return Font(name="Calibri", bold=True, size=9, color=_TEXT)

def _dfont() -> Font:
    return Font(name="Calibri", size=9, color=_TEXT)

def _mfont() -> Font:
    return Font(name="Calibri", size=9, color=_MUTED)

# ── Quarter utilities ─────────────────────────────────────────────────────────

def get_available_quarters(pnl: pd.Series) -> list[dict]:
    """
    Return completed calendar quarters present in pnl, newest first.

    A quarter is included when its end date is strictly before today
    and the quarter overlaps the pnl date range.

    Returns
    -------
    list of dict with keys: label (str), start (date), end (date).
    """
    today = date.today()
    pnl_min = pnl.index.min().date()
    pnl_max = pnl.index.max().date()

    month_abbr = {
        1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr",
        5: "May", 6: "Jun", 7: "Jul", 8: "Aug",
        9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
    }

    results: list[dict] = []
    for period in pd.period_range(pnl_min, today, freq="Q"):
        q_start = period.start_time.date()
        q_end   = period.end_time.date()

        if q_end >= today:          # incomplete quarter
            continue
        if q_end < pnl_min:        # entirely before data
            continue
        if q_start > pnl_max:      # entirely after data
            continue

        q_num = (q_start.month - 1) // 3 + 1
        label = (
            f"Q{q_num} {q_start.year} "
            f"({month_abbr[q_start.month]} – {month_abbr[q_end.month]} {q_start.year})"
        )
        results.append({"label": label, "start": q_start, "end": q_end})

    return list(reversed(results))


# ── Raw-data builder ──────────────────────────────────────────────────────────

def _populate_raw_data(wb: openpyxl.Workbook, quarter: dict, p1: dict, p2: dict,
                       alert_history: dict) -> None:
    """Write the selected quarter's daily rows into the Raw_Data table."""
    # Build a full-history regime series and forward-fill
    regime_series = pd.Series({d: v["regime"] for d, v in alert_history.items()})
    regime_series.index = pd.to_datetime(regime_series.index)
    regime_series = regime_series.sort_index()
    all_bdays = pd.date_range(regime_series.index.min(), regime_series.index.max(), freq="B")
    regime_full = regime_series.reindex(all_bdays).ffill().bfill()

    q_start = pd.Timestamp(quarter["start"])
    q_end   = pd.Timestamp(quarter["end"])

    hc_pnl = p1["pnl"]
    lc_pnl = p2["pnl"]

    # Rolling 252-day 5th-percentile VaR (computed on full history for accuracy)
    def _var95(series: pd.Series) -> pd.Series:
        return series.rolling(252, min_periods=30).quantile(0.05).abs()

    hc_var = _var95(hc_pnl)
    lc_var = _var95(lc_pnl)

    # NAV indexed to 1000 at first trading day of the quarter
    def _nav(series: pd.Series, start: pd.Timestamp) -> pd.Series:
        sub = series[series.index >= start]
        return (1 + sub).cumprod() * 1000

    hc_nav = _nav(hc_pnl, q_start)
    lc_nav = _nav(lc_pnl, q_start)

    hc_sub = hc_pnl[(hc_pnl.index >= q_start) & (hc_pnl.index <= q_end)]
    lc_sub = lc_pnl[(lc_pnl.index >= q_start) & (lc_pnl.index <= q_end)]

    rows: list[dict] = []
    for dt in sorted(set(hc_sub.index) | set(lc_sub.index)):
        regime = regime_full.get(dt, "Normal")
        for ft, src, var_s, nav_s in [
            ("Hard Currency", hc_sub, hc_var, hc_nav),
            ("Local Currency", lc_sub, lc_var, lc_nav),
        ]:
            if dt not in src.index:
                continue
            pnl_val = float(round(src[dt], 6))
            nav_val = float(round(nav_s.get(dt, 1000.0), 2))
            var_raw = var_s.get(dt, np.nan)
            var_val = float(round(var_raw, 6)) if not pd.isna(var_raw) else 0.005
            rows.append({
                "Date":             dt.to_pydatetime(),
                "Fund_Type":        ft,
                "Daily_Return":     pnl_val,
                "Nav":              nav_val,
                "VaR_95_Estimate":  var_val,
                "Actual_PnL":       pnl_val,
                "Regime_State":     regime,
            })

    ws = wb["Raw_Data"]
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    COLS = ["Date", "Fund_Type", "Daily_Return", "Nav",
            "VaR_95_Estimate", "Actual_PnL", "Regime_State"]
    for r_idx, record in enumerate(rows, start=2):
        for c_idx, col in enumerate(COLS, start=1):
            ws.cell(row=r_idx, column=c_idx).value = record[col]

    for tbl in ws.tables.values():
        if tbl.name == "tbl_RawData":
            tbl.ref = f"A1:G{len(rows) + 1}"


# ── Sheet layout helpers ──────────────────────────────────────────────────────

def _banner(ws, row: int, text: str, n_cols: int) -> None:
    ws.cell(row=row, column=1).value = text
    ws.cell(row=row, column=1).font = _hfont()
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).fill = _hfill()
    if n_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)


def _section(ws, row: int, text: str, n_cols: int) -> None:
    ws.cell(row=row, column=1).value = text
    ws.cell(row=row, column=1).font = _sfont()
    ws.row_dimensions[row].height = 18
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).fill = _sfill()
    if n_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)


def _col_headers(ws, row: int, headers: list[str]) -> None:
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = _cfill()
        cell.font = _cfont()
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30


def _data_cell(ws, row: int, col: int, value, alt: bool,
               fmt: str | None = None, align: str = "left") -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _rowfill(alt)
    cell.font = _dfont()
    cell.border = _border()
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt


def _fmt_val(val, fmt_str: str) -> float | str:
    """Return val as-is (let Excel format it) if numeric, else 'N/A'."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return "N/A"
    return val


# ── Risk Summary sheet ────────────────────────────────────────────────────────

def _write_risk_summary(wb: openpyxl.Workbook, rs1: dict, rs2: dict,
                         pn1: str, pn2: str, quarter_label: str) -> None:
    if "Risk_Summary" in wb.sheetnames:
        del wb["Risk_Summary"]
    ws = wb.create_sheet("Risk_Summary")
    wb.move_sheet("Risk_Summary", offset=-(len(wb.sheetnames) - 4))

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22

    _banner(ws, 1, f"EM FIXED INCOME — RISK SUMMARY   |   {quarter_label}", 3)
    _col_headers(ws, 2, ["METRIC", pn1, pn2])

    def _row(r, label, v1, v2, fmt=None, alt=False):
        _data_cell(ws, r, 1, label, alt)
        _data_cell(ws, r, 2, _fmt_val(v1, fmt), alt, fmt, align="center")
        _data_cell(ws, r, 3, _fmt_val(v2, fmt), alt, fmt, align="center")

    _section(ws, 3, "RETURN METRICS", 3)
    _row(4, "Cumulative Log Return (%)",   rs1["cum_log"],  rs2["cum_log"],  "0.00", alt=False)
    _row(5, "Annualised Return (%)",        rs1["ann_ret"],  rs2["ann_ret"],  "0.00", alt=True)
    _row(6, "Carry — Wtd Avg Yield (%)",   rs1["carry"],    rs2["carry"],    "0.00", alt=False)
    _row(7, "Roll-Down Return (est. %)",   rs1["rolldown"], rs2["rolldown"], "0.00", alt=True)

    _section(ws, 8, "RISK & RATIO METRICS", 3)
    _row(9,  "Annualised Volatility (%)",      rs1["ann_vol"],     rs2["ann_vol"],     "0.00",  alt=False)
    _row(10, "Maximum Drawdown (%)",           rs1["max_dd"],      rs2["max_dd"],      "0.00",  alt=True)
    _row(11, "Sharpe Ratio (rf = €STR)",       rs1["sharpe"],      rs2["sharpe"],      "0.00",  alt=False)
    _row(12, "Sortino Ratio (rf = €STR)",      rs1["sortino"],     rs2["sortino"],     "0.00",  alt=True)
    _row(13, "Sharpe Ratio (rf = 0, ref)",     rs1["sharpe_zero"], rs2["sharpe_zero"], "0.00",  alt=False)
    _row(14, "Sortino Ratio (MAR = 0, ref)",   rs1["sortino_zero"],rs2["sortino_zero"],"0.00",  alt=True)
    _row(15, "Calmar Ratio",                   rs1["calmar"],      rs2["calmar"],      "0.00",  alt=False)

    _section(ws, 16, "BOND ANALYTICS", 3)
    _row(17, "AUM (EUR)",                       rs1["aum"],       rs2["aum"],       '#,##0',   alt=False)
    _row(18, "Modified Duration (yrs)",         rs1["mod_dur"],   rs2["mod_dur"],   "0.00",    alt=True)
    _row(19, "DV01 (% of NAV per 1bp)",         rs1["dv01"],      rs2["dv01"],      "0.0000",  alt=False)
    _row(20, "DV01 (EUR per 1bp parallel)",     rs1["dv01_eur"],  rs2["dv01_eur"],  '#,##0',   alt=True)
    _row(21, "Convexity (yrs²)",                rs1["convexity"], rs2["convexity"], "0.0",     alt=False)
    _row(22, "YTM — Wtd Avg Benchmark (%)",     rs1["ytm"],       rs2["ytm"],       "0.00",    alt=True)
    _row(23, "Yield Curve Slope (long−short,%)",rs1["yc_slope"],  rs2["yc_slope"],  "0.00",    alt=False)


# ── VaR Analysis sheet ────────────────────────────────────────────────────────

def _write_var_analysis(wb: openpyxl.Workbook, rs1: dict, rs2: dict,
                         pn1: str, pn2: str, quarter_label: str) -> None:
    if "VaR_Analysis" in wb.sheetnames:
        del wb["VaR_Analysis"]
    ws = wb.create_sheet("VaR_Analysis")
    wb.move_sheet("VaR_Analysis", offset=-(len(wb.sheetnames) - 5))

    for col, w in zip("ABCDEFG", [28, 14, 20, 20, 18, 18, 18]):
        ws.column_dimensions[col].width = w

    _banner(ws, 1, f"EM FIXED INCOME — VAR & CVAR ANALYSIS   |   {quarter_label}", 7)

    # % section
    _section(ws, 2, "VAR & CVAR AS % OF PORTFOLIO NAV", 7)
    _col_headers(ws, 3, ["Portfolio", "Confidence",
                          "Param VaR (%)", "Param CVaR (%)",
                          "Hist VaR (%)", "Hist CVaR (%)", "MC VaR (%)"])
    r = 4
    for pname, rs in [(pn1, rs1), (pn2, rs2)]:
        for vr in rs["var_rows"]:
            alt = (r % 2 == 0)
            _data_cell(ws, r, 1, pname,             alt, align="left")
            _data_cell(ws, r, 2, vr["Confidence"],  alt, align="center")
            _data_cell(ws, r, 3, vr["Param VaR (%)"],  alt, "0.0000", "center")
            _data_cell(ws, r, 4, vr["Param CVaR (%)"], alt, "0.0000", "center")
            _data_cell(ws, r, 5, vr["Hist VaR (%)"],   alt, "0.0000", "center")
            _data_cell(ws, r, 6, vr["Hist CVaR (%)"],  alt, "0.0000", "center")
            _data_cell(ws, r, 7, vr["MC VaR (%)"],     alt, "0.0000", "center")
            r += 1

    r += 1  # blank row

    # EUR section
    _section(ws, r, "VAR & CVAR IN EUR (BASED ON AUM)", 7)
    r += 1
    _col_headers(ws, r, ["Portfolio", "Confidence",
                          "Param VaR (EUR)", "Param CVaR (EUR)",
                          "Hist VaR (EUR)", "MC VaR (EUR)", ""])
    r += 1
    for pname, rs in [(pn1, rs1), (pn2, rs2)]:
        for vr in rs["var_rows_eur"]:
            alt = (r % 2 == 0)
            _data_cell(ws, r, 1, pname,                      alt, align="left")
            _data_cell(ws, r, 2, vr["Confidence"],           alt, align="center")
            _data_cell(ws, r, 3, vr["Param VaR (EUR)"],      alt, '#,##0', "center")
            _data_cell(ws, r, 4, vr["Param CVaR (EUR)"],     alt, '#,##0', "center")
            _data_cell(ws, r, 5, vr["Hist VaR (EUR)"],       alt, '#,##0', "center")
            _data_cell(ws, r, 6, vr["MC VaR (EUR)"],         alt, '#,##0', "center")
            ws.cell(row=r, column=7).fill = _rowfill(alt)
            r += 1


# ── KRD sheet ─────────────────────────────────────────────────────────────────

def _write_krd(wb: openpyxl.Workbook, rs1: dict, rs2: dict,
               pn1: str, pn2: str, quarter_label: str) -> None:
    if "KRD" in wb.sheetnames:
        del wb["KRD"]
    ws = wb.create_sheet("KRD")
    wb.move_sheet("KRD", offset=-(len(wb.sheetnames) - 6))

    for col, w in zip("ABCD", [22, 22, 26, 26]):
        ws.column_dimensions[col].width = w

    _banner(ws, 1, f"EM FIXED INCOME — KEY-RATE DURATION   |   {quarter_label}", 4)

    _section(ws, 2, "KEY-RATE DURATION BY COUNTRY (yrs)", 4)
    _col_headers(ws, 3, ["Country", "MD (par bond, yrs)",
                          f"KRD — {pn1} (yrs)", f"KRD — {pn2} (yrs)"])

    FLAG = {"Brazil": "\U0001f1e7\U0001f1f7", "Mexico": "\U0001f1f2\U0001f1fd",
            "South Africa": "\U0001f1ff\U0001f1e6", "Poland": "\U0001f1f5\U0001f1f1",
            "Colombia": "\U0001f1e8\U0001f1f4", "Hungary": "\U0001f1ed\U0001f1fa",
            "Romania": "\U0001f1f7\U0001f1f4"}

    all_countries = sorted(set(rs1["krd"]) | set(rs2["krd"]))
    for i, c in enumerate(all_countries):
        r = 4 + i
        alt = (i % 2 == 1)
        label = f"{FLAG.get(c, '')} {c}"
        md = rs1["md_by_c"].get(c, 0)
        _data_cell(ws, r, 1, label,                 alt)
        _data_cell(ws, r, 2, md,                    alt, "0.000", "center")
        _data_cell(ws, r, 3, rs1["krd"].get(c, 0), alt, "0.0000", "center")
        _data_cell(ws, r, 4, rs2["krd"].get(c, 0), alt, "0.0000", "center")

    snap_start = 4 + len(all_countries) + 2
    _section(ws, snap_start, "LATEST BENCHMARK YIELD SNAPSHOT (%)", 4)
    _col_headers(ws, snap_start + 1, ["Country", "Latest Yield (%)", "As of", ""])
    for i, c in enumerate(all_countries):
        r = snap_start + 2 + i
        alt = (i % 2 == 1)
        label = f"{FLAG.get(c, '')} {c}"
        y_val = rs1["c_vals"].get(c) or rs2["c_vals"].get(c)
        _data_cell(ws, r, 1, label,                           alt)
        _data_cell(ws, r, 2, y_val if y_val is not None else "N/A", alt, "0.00", "center")
        _data_cell(ws, r, 3, "Latest available",              alt, align="center")
        ws.cell(row=r, column=4).fill = _rowfill(alt)


# ── Main entry point ──────────────────────────────────────────────────────────

def generate_quarterly_report(quarter: dict) -> bytes:
    """
    Build a quarterly Excel report for the given quarter dict.

    Parameters
    ----------
    quarter : dict
        One entry from get_available_quarters() —
        must have keys: label (str), start (date), end (date).

    Returns
    -------
    bytes — the xlsx file contents, ready for st.download_button.
    """
    from src.data.var_artifacts import load_alert_history as _load_alerts
    from src.data_loader import (
        load_config,
        load_all_countries_combined,
        load_country_yields,
    )
    from src.risk_free import load_risk_free_rates
    from src.services.portfolios import build_portfolio_views
    from src.services.risk_stats import compute_risk_stats

    warnings.filterwarnings("ignore")

    cfg = load_config()
    portfolio_results = build_portfolio_views()
    p1, p2 = portfolio_results[0], portfolio_results[1]

    alert_history = _load_alerts(DATA_OUT) or {}

    # Yield levels for risk-stat computation
    all_countries = (
        cfg["countries"]["local_currency"] + cfg["countries"]["hard_currency"]
    )
    excluded = cfg.get("excluded_series", {})
    yield_levels: dict = {}
    for country in all_countries:
        try:
            df = load_country_yields(country, "data/raw")
            excl = excluded.get(country, [])
            df = df.drop(columns=[c for c in excl if c in df.columns])
            yield_levels[country] = df
        except Exception:
            pass

    rf_data = None
    try:
        fred_cfg = cfg.get("fred", {})
        key = open(fred_cfg.get("key_path", "private/fred_key.txt")).read().strip()
        rf_data = load_risk_free_rates(
            fred_cfg.get("output_path", "data/output/risk_free_rates.csv"),
            fred_api_key=key,
        )
    except Exception:
        pass

    rs1 = compute_risk_stats(p1["def"], p1["pnl"], yield_levels, rf_data)
    rs2 = compute_risk_stats(p2["def"], p2["pnl"], yield_levels, rf_data)
    pn1, pn2 = p1["def"]["name"], p2["def"]["name"]

    # Load and populate workbook
    wb = openpyxl.load_workbook(TEMPLATE_PATH)

    _populate_raw_data(wb, quarter, p1, p2, alert_history)

    ws_perf = wb["Performance_Report"]
    ws_perf["D5"] = datetime.combine(quarter["start"], datetime.min.time())
    ws_perf["D6"] = datetime.combine(quarter["end"],   datetime.min.time())

    lbl = quarter["label"]
    _write_risk_summary(wb, rs1, rs2, pn1, pn2, lbl)
    _write_var_analysis(wb, rs1, rs2, pn1, pn2, lbl)
    _write_krd(wb, rs1, rs2, pn1, pn2, lbl)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
