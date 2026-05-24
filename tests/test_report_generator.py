"""Tests for src/report_generator.py."""
import io
from datetime import date

import numpy as np
import pandas as pd
import openpyxl
import pytest

from src.report_generator import get_available_quarters, generate_quarterly_report


def _make_pnl(start: str, end: str) -> pd.Series:
    rng = np.random.default_rng(42)
    idx = pd.bdate_range(start, end)
    return pd.Series(rng.normal(0.0002, 0.003, len(idx)), index=idx)


def test_returns_nonempty_list():
    pnl = _make_pnl("2024-01-02", "2026-04-30")
    quarters = get_available_quarters(pnl)
    assert isinstance(quarters, list)
    assert len(quarters) > 0


def test_newest_first():
    pnl = _make_pnl("2024-01-02", "2026-04-30")
    quarters = get_available_quarters(pnl)
    for i in range(len(quarters) - 1):
        assert quarters[i]["start"] > quarters[i + 1]["start"]


def test_all_quarters_completed():
    pnl = _make_pnl("2024-01-02", "2026-04-30")
    today = date.today()
    for q in get_available_quarters(pnl):
        assert q["end"] < today


def test_label_format():
    pnl = _make_pnl("2025-01-02", "2025-06-30")
    quarters = get_available_quarters(pnl)
    labels = [q["label"] for q in quarters]
    assert any("Q1 2025" in lbl for lbl in labels)
    assert any("Q2 2025" in lbl for lbl in labels)


def test_required_keys():
    pnl = _make_pnl("2024-01-02", "2026-04-30")
    q = get_available_quarters(pnl)[0]
    assert "label" in q
    assert "start" in q
    assert "end" in q
    assert isinstance(q["start"], date)
    assert isinstance(q["end"], date)


def test_generate_quarterly_report_valid_xlsx():
    """Integration test — requires Report_Template.xlsx and data/output/ on disk."""
    q = {"label": "Q1 2026 (Jan – Mar 2026)", "start": date(2026, 1, 1), "end": date(2026, 3, 31)}
    result = generate_quarterly_report(q)
    assert isinstance(result, bytes)
    assert len(result) > 5_000

    wb = openpyxl.load_workbook(io.BytesIO(result))
    for sheet in ["Raw_Data", "Performance_Report", "Alerts_Review",
                  "Risk_Summary", "VaR_Analysis", "KRD"]:
        assert sheet in wb.sheetnames, f"Missing sheet: {sheet}"

    # Raw_Data must have data rows for Q1 2026
    ws = wb["Raw_Data"]
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None]
    assert len(data_rows) > 0

    # Report dates should reflect the quarter
    ws_perf = wb["Performance_Report"]
    assert ws_perf["D5"].value is not None   # Report_Start_Date set
    assert ws_perf["D6"].value is not None   # Report_End_Date set
